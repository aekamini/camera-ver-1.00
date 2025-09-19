
def normalize_similarity(score, low_baseline=50.0, high_baseline=100.0):
    """
    نرمال‌سازی شباهت: هر چیزی زیر low_baseline → 0%
    هر چیزی بالای high_baseline → 100%
    بین این دو خطی مپ می‌شود.
    """
    if score <= low_baseline:
        return 0.0
    if score >= high_baseline:
        return 100.0
    return (score - low_baseline) / (high_baseline - low_baseline) * 100.0


import cv2
import numpy as np
from PIL import Image
from skimage.metrics import structural_similarity as ssim
from skimage.exposure import match_histograms
import imagehash

def phash_similarity(pil1, pil2):
    try:
        h1 = imagehash.phash(pil1)
        h2 = imagehash.phash(pil2)
        dist = h1 - h2
        sim = 100.0 - (dist / 64.0 * 100.0)
        return float(sim), int(dist)
    except Exception:
        return None, None

def align_images_orb(ref_gray, test_gray):
    orb = cv2.ORB_create(2000)
    kp1, des1 = orb.detectAndCompute(ref_gray, None)
    kp2, des2 = orb.detectAndCompute(test_gray, None)
    if des1 is None or des2 is None or len(kp1) < 4 or len(kp2) < 4:
        return None, 0.0
    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
    matches = bf.match(des1, des2)
    matches = sorted(matches, key=lambda x: x.distance)
    if len(matches) < 4:
        return None, 0.0
    src_pts = np.float32([kp1[m.queryIdx].pt for m in matches]).reshape(-1,1,2)
    dst_pts = np.float32([kp2[m.trainIdx].pt for m in matches]).reshape(-1,1,2)
    H, mask = cv2.findHomography(dst_pts, src_pts, cv2.RANSAC, 5.0)
    good_ratio = float(mask.sum()) / float(len(matches)) if mask is not None else 0.0
    return H, good_ratio



def compare_images_pipeline(test_pil, ref_pil, polygon_points=None, target_size=(300,300), do_align=True, app=None):
    ref_rgb = np.array(ref_pil.convert("RGB"))
    test_rgb = np.array(test_pil.convert("RGB"))

    ref_gray_full = cv2.cvtColor(ref_rgb, cv2.COLOR_RGB2GRAY)
    test_gray_full = cv2.cvtColor(test_rgb, cv2.COLOR_RGB2GRAY)

    H = None
    orb_ratio = 0.0
    if do_align:
        H, orb_ratio = align_images_orb(ref_gray_full, test_gray_full)
        if H is not None:
            h_ref, w_ref = ref_gray_full.shape
            test_rgb = cv2.warpPerspective(test_rgb, H, (w_ref, h_ref))
            test_gray_full = cv2.cvtColor(test_rgb, cv2.COLOR_RGB2GRAY)

    ref_resized = cv2.resize(ref_rgb, target_size, interpolation=cv2.INTER_AREA)
    test_resized = cv2.resize(test_rgb, target_size, interpolation=cv2.INTER_AREA)

    try:
        matched = match_histograms(test_resized, ref_resized, multichannel=True)
        test_resized = np.clip(matched, 0, 255).astype(np.uint8)
    except Exception:
        pass

    ref_gray = cv2.cvtColor(ref_resized, cv2.COLOR_RGB2GRAY)
    test_gray = cv2.cvtColor(test_resized, cv2.COLOR_RGB2GRAY)

    try:
        ssim_val = ssim(ref_gray, test_gray, data_range=255)
    except Exception:
        ssim_val = 0.0

    diffs = np.abs(ref_gray.astype(int) - test_gray.astype(int))
    diff_mean = diffs.mean()
    diff_pct = 100.0 - (diff_mean / 255.0 * 100.0)

    phash_sim, phash_dist = phash_similarity(Image.fromarray(test_resized), Image.fromarray(ref_resized))
    phash_sim_val = phash_sim if phash_sim is not None else 0.0

    orb_score = float(orb_ratio) * 100.0

    try:
        hsv_ref = cv2.cvtColor(ref_resized, cv2.COLOR_RGB2HSV)
        hsv_test = cv2.cvtColor(test_resized, cv2.COLOR_RGB2HSV)
        hist_r_h = cv2.calcHist([hsv_ref], [0], None, [50], [0, 180])
        hist_t_h = cv2.calcHist([hsv_test], [0], None, [50], [0, 180])
        cv2.normalize(hist_r_h, hist_r_h)
        cv2.normalize(hist_t_h, hist_t_h)
        corr_h = cv2.compareHist(hist_r_h, hist_t_h, cv2.HISTCMP_CORREL)
        hist_r_s = cv2.calcHist([hsv_ref], [1], None, [50], [0, 256])
        hist_t_s = cv2.calcHist([hsv_test], [1], None, [50], [0, 256])
        cv2.normalize(hist_r_s, hist_r_s)
        cv2.normalize(hist_t_s, hist_t_s)
        corr_s = cv2.compareHist(hist_r_s, hist_t_s, cv2.HISTCMP_CORREL)
        hist_score = ((max(-1.0, min(1.0, corr_h)) + 1.0) / 2.0 * 100.0 + (max(-1.0, min(1.0, corr_s)) + 1.0) / 2.0 * 100.0) / 2.0
    except Exception:
        hist_score = 50.0

    w_ssim, w_diff, w_phash, w_orb, w_hist = 0.15, 0.25, 0.35, 0.10, 0.15

    combined = (w_ssim * (ssim_val * 100.0) +
                w_diff * diff_pct +
                w_phash * phash_sim_val +
                w_orb * orb_score +
                w_hist * hist_score)

    # گرفتن تنظیمات از UI اگر موجود بود
    if app and hasattr(app, "sliders_win"):
        low_baseline = app.sliders_win.sensitivity_var.get()
        phash_thr = app.sliders_win.phash_thr_var.get()
        gamma = app.sliders_win.gamma_var.get() / 10.0
    else:
        low_baseline = 55
        phash_thr = 40
        gamma = 1.7

    if phash_sim_val < phash_thr:
        combined = min(combined, phash_sim_val)
    if orb_score < 5.0 and phash_sim_val < phash_thr:
        combined = combined * 0.5

    def normalize_similarity_local(score, low_baseline=55.0, high_baseline=100.0):
        if score <= low_baseline:
            return 0.0
        if score >= high_baseline:
            return 100.0
        return (score - low_baseline) / (high_baseline - low_baseline) * 100.0

    normalized = normalize_similarity_local(combined, low_baseline=low_baseline, high_baseline=100.0)
    final_score = 100.0 * ((normalized / 100.0) ** gamma) if normalized > 0 else 0.0

    result = {
        "ssim": float(ssim_val) * 100.0,
        "diff_pct": float(diff_pct),
        "phash_sim": float(phash_sim_val),
        "phash_dist": phash_dist,
        "orb_score": orb_score,
        "hist_score": float(hist_score),
        "combined_score": float(final_score)
    }
    return result
import tkinter as tk
from tkinter import simpledialog, messagebox
from tkinter import ttk
import os, json, datetime, threading
from PIL import Image, ImageTk
import cv2

import numpy as np
from skimage.metrics import structural_similarity as ssim
import sqlite3
APP_FOLDER = "app_data_full"
SETTINGS_FILE = os.path.join(APP_FOLDER, "settings.json")
LOG_FILE = os.path.join(APP_FOLDER, "log.txt")
REFERENCE_FILE = os.path.join(APP_FOLDER, "reference.png")
os.makedirs(APP_FOLDER, exist_ok=True)
DEFAULT_PASSWORD = "1234"
# --- Database helper: initialize and save results ---
DB_FILE = os.path.join(APP_FOLDER, "results.db")
from openpyxl import Workbook, load_workbook
import os
EXCEL_PATH = os.path.join(APP_FOLDER, "results_live.xlsx")
def append_to_excel(data):
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Filename", "Timestamp", "SSIM", "Diff %", "pHash", "ORB", "Combined", "Result"])
    else:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    ws.append(data)
    wb.save(EXCEL_PATH)
def init_db():
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT,
                datetime TEXT,
                similarity REAL,
                result TEXT
            )
        """)
   
        append_to_excel([filename, timestamp,
                         result_dict["ssim"], result_dict["diff_pct"],
                         result_dict["phash_sim"], result_dict["orb_score"],
                         result_dict["combined_score"], result_text])
        conn.close()
    except Exception:
        pass
def save_result_db(filename, similarity, result):
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute("INSERT INTO results (filename, datetime, similarity, result) VALUES (?, ?, ?, ?)", 
                  (filename, now, float(similarity), result))

        append_to_excel([filename, timestamp,
                         result_dict["ssim"], result_dict["diff_pct"],
                         result_dict["phash_sim"], result_dict["orb_score"],
                         result_dict["combined_score"], result_text])
        conn.close()
    except Exception:
        pass

# initialize DB on import
init_db()


def log_message(text, app=None):
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {text}"
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\\n")
    except Exception:
        pass
    if app and hasattr(app, "msg_viewer") and hasattr(app.msg_viewer, "append_message"):
        app.msg_viewer.append_message(line)


class DraggableButton(tk.Button):
    def __init__(self, master, text, command=None, init_pos=(10, 10), style=None, **kwargs):
        kwargs = kwargs or {}
        if style:
            kwargs.update(style)
        super().__init__(master, text=text, command=command, **kwargs)
        self.locked = False  # 🔒 قفل پیش‌فرض خاموش
        self.place(x=init_pos[0], y=init_pos[1])
        self.bind("<Button-1>", self._start_drag)
        self.bind("<B1-Motion>", self._on_drag)
        self._drag_start = (0, 0)
        self.command = command

    def _start_drag(self, event):
        if self.locked:
            return
        self._drag_start = (event.x, event.y)

    def _on_drag(self, event):
        if self.locked:
            return
        dx = event.x - self._drag_start[0]
        dy = event.y - self._drag_start[1]
        info = self.place_info()
        try:
            new_x = int(float(info.get("x", 0))) + dx
            new_y = int(float(info.get("y", 0))) + dy
        except Exception:
            new_x, new_y = dx, dy
        self.place(x=new_x, y=new_y)
        self.lift()


class RoundedFrame(tk.Frame):
    def __init__(self, master, radius=15, bg="#F9FAFB", **kwargs):
        super().__init__(master, bg=bg, **kwargs)
        self.config(highlightthickness=1, highlightbackground="#CBD5E1")


class MDIChild(RoundedFrame):
    def __init__(self, master, title, key, default_geom, settings,
                 with_buttons=False, with_sliders=False, with_text=False,
                 with_video=False, with_reference=False, with_test=False, fixed=False):
        super().__init__(master, bg="#F9FAFB")
        self.master_app = master.master
        self.key = key
        self.settings = settings
        self.title = title
        self.video_cap = None
        self.fixed = fixed  # store fixed flag

        if key in settings.get("windows", {}):
            x, y, w, h = settings["windows"][key]
        else:
            x, y, w, h = default_geom
        self.place(x=x, y=y, width=w, height=h)

        if not self.fixed:
            self.header = tk.Frame(self, bg="#1E3A8A", height=28, cursor="fleur")
            self.header.pack(fill="x", side="top")
            self.lbl = tk.Label(self.header, text=title, bg="#1E3A8A", fg="white", font=("Arial", 10, "bold"))
            self.lbl.pack(side="left", padx=8)
            self.header.bind("<Button-1>", self._start_move)
            self.header.bind("<B1-Motion>", self._on_move)

        self.body = tk.Frame(self, bg="#F9FAFB")
        self.body.pack(fill="both", expand=True)

        self.text_area = None
        self.video_label = None
        self.canvas = None
        self.sliders_frame = None
        self.run_btn = None
        self.exit_btn = None

        if with_buttons and key == "settings":
            self._make_settings_buttons()

        if with_sliders:
            self._make_sliders()

        if with_text:
            self.text_area = tk.Text(self.body, bg="black", fg="#00FF7F", font=("Consolas", 11), wrap="word")
            self.text_area.pack(fill="both", expand=True)

        if with_video:
            self.video_label = tk.Label(self.body, bg="black")
            self.video_label.pack(fill="both", expand=True)

        if with_reference:
            self.canvas = tk.Canvas(self.body, bg="black", highlightthickness=0)
            self.canvas.pack(fill="both", expand=True)

        if with_test:
            self.text_area = tk.Text(self.body, bg="white", fg="black", font=("Arial", 11), wrap="word")
            self.text_area.pack(fill="both", expand=True)

        if not self.fixed:
            self.resize_handle = tk.Frame(self, bg="gray", cursor="size_nw_se")
            self.resize_handle.place(relx=1.0, rely=1.0, anchor="se", width=16, height=16)
            self.resize_handle.lift()
            self.resize_handle.bind("<Button-1>", self._start_resize)
            self.resize_handle.bind("<B1-Motion>", self._on_resize)

        # When the canvas resizes, keep the reference image updated if present
        self._ref_img_pil = None
        if self.canvas:
            self.canvas.bind("<Configure>", self._on_canvas_configure)

    def _on_canvas_configure(self, event):
        # redraw reference if we have one
        if getattr(self, "_ref_img_pil", None):
            self.show_reference(self._ref_img_pil)

    def _make_settings_buttons(self):
        btn_positions = self.settings.get("button_positions", {})
        style = dict(bg="#2563EB", fg="white", font=("Arial", 10, "bold"),
                     activebackground="#3B82F6", activeforeground="white",
                     width=12, height=2, relief="raised", bd=2)

        def make_btn(name, requires_pwd=False, default=(10, 10)):
            pos = btn_positions.get(name, default)
            def handler(n=name, pwd=requires_pwd):
                if pwd:
                    pw = simpledialog.askstring("Password", "Enter password:", show="*")
                    if pw != self.settings.get("admin_password", DEFAULT_PASSWORD):
                        messagebox.showerror("Error", "Wrong password!")
                        return
                app = self.master_app
                if n == "Reference":
                    if app.live_win:
                        img = app.live_win.capture_frame()
                        if img:
                            # save to file (overwrite if exists)
                            try:
                                img.save(REFERENCE_FILE)
                                log_message(f"Saved reference to {REFERENCE_FILE}", app)
                            except Exception as e:
                                log_message(f"Failed to save reference: {e}", app)
                                messagebox.showerror("Error", f"Failed to save reference:\\n{e}")
                                return
                            # show in reference window
                            app.ref_win.show_reference(img)
                            app.test_win.append_message_simple("Reference captured and saved")
                        else:
                            messagebox.showerror("Error", "No frame from Live")
                elif n == "Polygon":
                    if app.ref_win:
                        app.ref_win.start_polygon_mode()
                        log_message("Polygon drawing started", app)
                        app.test_win.append_message_simple("Polygon drawing started")
                elif n == "Reset Polygon":
                    if app.ref_win:
                        app.ref_win.clear_polygon()
                        log_message("Polygon reset", app)
                        app.test_win.append_message_simple("Polygon reset")
                elif n == "Admin Panel":
                    self._open_admin_panel()
            return DraggableButton(self.body, name, command=handler, init_pos=(pos[0], pos[1]), style=style)

        self.ref_btn = make_btn("Reference", requires_pwd=True, default=(20, 40))
        self.poly_btn = make_btn("Polygon", requires_pwd=True, default=(160, 40))
        self.reset_btn = make_btn("Reset Polygon", requires_pwd=True, default=(300, 40))
        self.admin_btn = make_btn("Admin Panel", requires_pwd=True, default=(20, 100))
        # Device selection (Webcam / Capture card)
        # --- Device selection (draggable) ---
        btn_positions = self.settings.get("button_positions", {})
        pos_label = btn_positions.get("Select Device:", (20, 160))   # اصلاح برای ذخیره درست
        pos_combo = btn_positions.get("Device Combo", (120, 160))

        # Draggable Label
        self.device_label = DraggableButton(
            self.body, "Select Device:", command=None,
            init_pos=(pos_label[0], pos_label[1]),
            style=dict(bg="#F9FAFB", fg="black", relief="flat", width=15, height=1)
        )

        # Frame + Combobox
        frame_combo = tk.Frame(self.body, bg="#F9FAFB")
        self.device_var = tk.StringVar(value=self.settings.get("device", "Capture Card"))
        device_combo = ttk.Combobox(frame_combo, textvariable=self.device_var,
                                    values=["Capture Card", "Webcam"], state="readonly", width=15)
        device_combo.pack()

        # Draggable wrapper برای Combobox
        self.device_combo = DraggableButton(
            self.body, "", command=None,
            init_pos=(pos_combo[0], pos_combo[1]),
            style=dict(width=0, height=0, relief="flat", bg="#F9FAFB")
        )
        frame_combo.place(in_=self.device_combo, x=0, y=0)

        def save_device_choice(event=None):
            choice = self.device_var.get()
            self.settings["device"] = choice
            try:
                with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.settings, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            if hasattr(self.master_app, "change_device"):
                self.master_app.change_device(choice)

        device_combo.bind("<<ComboboxSelected>>", save_device_choice)

        # --- Metric Toggles (Row under Admin, draggable & persisted) ---
        if "enabled_metrics" not in self.settings:
            self.settings["enabled_metrics"] = {"ssim": True, "diff": True, "phash": True, "orb": True, "hist": True}
        if "metric_order" not in self.settings:
            self.settings["metric_order"] = ["ssim", "diff", "phash", "orb", "hist"]

        # Frame for metric buttons placed below Admin Panel
        self.metric_frame = tk.Frame(self.body, bg="#F9FAFB")
        self.metric_frame.pack(pady=10, anchor="w")

        self.metric_buttons = {}
        for m in self.settings["metric_order"]:
            self._create_metric_button(m)




    # ---------- Metric button helpers (row layout, drag & drop, persistence) ----------
    def _create_metric_button(self, name):
        btn = tk.Button(
            self.metric_frame,
            text=f"{name.upper()} ON" if self.settings["enabled_metrics"].get(name, True) else f"{name.upper()} OFF",
            bg="green" if self.settings["enabled_metrics"].get(name, True) else "red",
            fg="white", width=12, height=1
        )
        btn.pack(pady=2, anchor="w")
        btn.bind("<Button-1>", lambda e, n=name: self._on_drag_start(e, n))
        btn.bind("<B1-Motion>", self._on_drag_motion)
        btn.bind("<ButtonRelease-1>", lambda e, n=name: self._on_drag_release(e, n))
        btn.config(command=lambda n=name, b=btn: self.toggle_metric(n, b))
        self.metric_buttons[name] = btn

    def _on_drag_start(self, event, name):
        if self.locked:
            return
        self._dragging_name = name
        self._drag_start_index = self.settings["metric_order"].index(name)

    def _on_drag_motion(self, event):
        if self.locked:
            return
        pass

    def _on_drag_release(self, event, name):
        if self.locked:
            return
        if not hasattr(self, "_dragging_name"):
            return
        old_idx = self._drag_start_index
        y = event.y_root - self.metric_frame.winfo_rooty()
        row_height = event.widget.winfo_height() + 4
        new_idx = int(y // row_height)
        order = self.settings["metric_order"]
        if 0 <= new_idx < len(order) and new_idx != old_idx:
            item = order.pop(old_idx)
            order.insert(new_idx, item)
            for w in self.metric_frame.winfo_children():
                w.destroy()
            self.metric_buttons = {}
            for m in order:
                self._create_metric_button(m)
            try:
                with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.settings, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
        self._dragging_name = None

    def toggle_metric(self, name, btn):
        current = self.settings["enabled_metrics"].get(name, True)
        self.settings["enabled_metrics"][name] = not current
        if self.settings["enabled_metrics"][name]:
            btn.config(bg="green", text=f"{name.upper()} ON")
        else:
            btn.config(bg="red", text=f"{name.upper()} OFF")
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(self.settings, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _make_sliders(self):
        self.sliders_frame = tk.Frame(self.body, bg="#F9FAFB")
        self.sliders_frame.pack(fill="both", expand=True)
        svals = self.settings.get("sliders", {"min": 0, "max": 100})
        self.min_var = tk.IntVar(value=int(svals.get("min", 0)))
        self.max_var = tk.IntVar(value=int(svals.get("max", 100)))
        
        # فریم بالایی برای Min/Max
        top_frame = tk.Frame(self.sliders_frame, bg="#F9FAFB")
        top_frame.pack(side="top", fill="x", pady=5)

        self.min_slider = tk.Scale(top_frame, from_=100, to=0, orient="vertical",
                                   variable=self.min_var, length=80, label="Min",
                                   bg="#F9FAFB", troughcolor="#93C5FD", highlightthickness=0)
        self.min_slider.pack(side="left", padx=20)

        self.max_slider = tk.Scale(top_frame, from_=100, to=0, orient="vertical",
                                   variable=self.max_var, length=80, label="Max",
                                   bg="#F9FAFB", troughcolor="#93C5FD", highlightthickness=0)
        self.max_slider.pack(side="left", padx=20)

        btn_positions = self.settings.get("button_positions", {})
        style = dict(bg="#2563EB", fg="white", font=("Arial", 10, "bold"),
                     activebackground="#3B82F6", activeforeground="white",
                     width=12, height=2, relief="raised", bd=2)

        pos_run = btn_positions.get("Export to Excel", (20, 100))
        def run_handler():
            log_message("Run Test executed", self.master_app)
            if self.master_app.test_win:
                self.master_app.test_win.append_message_simple("Run Test executed")

        # --- Export to Excel function ---
        from openpyxl import Workbook
        def export_to_excel():
            import sqlite3, os
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT id, filename, datetime, similarity, result FROM results ORDER BY id DESC")
            rows = c.fetchall()
            conn.close()
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Results"
            headers = ["ID", "Filename", "DateTime", "Similarity (%)", "Result"]
            ws.append(headers)
            for row in rows:
                ws.append(row)
            xlsx_path = os.path.join(APP_FOLDER, "test_results.xlsx")
            wb.save(xlsx_path)
            try:
                os.startfile(xlsx_path)
            except Exception:
                print("Excel file saved at:", xlsx_path)
            if self.master_app.test_win:
                self.master_app.test_win.append_message_simple(f"✅ Exported results to {xlsx_path}")

        self.export_btn = DraggableButton(self.body, "Export to Excel", command=export_to_excel, init_pos=(pos_run[0], pos_run[1]), style=style)

        

        # --- اضافه کردن اسلایدرهای جدید به صورت مرتب شده ---
        btn_positions = self.settings.get("button_positions", {})

        # فریم پایینی برای اسلایدرهای حساسیت، pHash و Gamma
        bottom_frame = tk.Frame(self.sliders_frame, bg="#F9FAFB")
        bottom_frame.pack(side="top", fill="both", expand=True)

        # Sensitivity Slider
        pos_sens = btn_positions.get("Sensitivity Slider", (20, 200))
        self.sensitivity_var = tk.IntVar(value=55)
        scale_sens = tk.Scale(bottom_frame, from_=30, to=70, orient="horizontal",
                              variable=self.sensitivity_var, length=200,
                              label="Sensitivity (Baseline)",
                              bg="#F9FAFB", troughcolor="#93C5FD", highlightthickness=0)
        scale_sens.pack(pady=5)

        # pHash Threshold Slider
        pos_phash = btn_positions.get("pHash Slider", (20, 300))
        self.phash_thr_var = tk.IntVar(value=40)
        scale_phash = tk.Scale(bottom_frame, from_=20, to=60, orient="horizontal",
                               variable=self.phash_thr_var, length=200,
                               label="pHash Threshold",
                               bg="#F9FAFB", troughcolor="#93C5FD", highlightthickness=0)
        scale_phash.pack(pady=5)

        # Gamma Slider
        pos_gamma = btn_positions.get("Gamma Slider", (20, 400))
        self.gamma_var = tk.IntVar(value=17)  # gamma *10
        scale_gamma = tk.Scale(bottom_frame, from_=10, to=30, orient="horizontal",
                               variable=self.gamma_var, length=200, resolution=1,
                               label="Gamma (x0.1)",
                               bg="#F9FAFB", troughcolor="#93C5FD", highlightthickness=0)
        scale_gamma.pack(pady=5)

        pos_start =  btn_positions.get("Start Test", (20, 140))
        
        def start_test_handler():
            app = self.master_app
            if app.live_win:
                img = app.live_win.capture_frame()
                if img is None:
                    if app.test_win:
                        app.test_win.append_message_simple("❌ No frame available from Live")
                    return

                fname = simpledialog.askstring("Save Test Image", "Enter file name (without extension):")
                if not fname or not fname.strip():
                    if app.test_win:
                        app.test_win.append_message_simple("❌ File name cannot be empty")
                    return
                fname = fname.strip() + ".png"
                fpath = os.path.join(APP_FOLDER, fname)

                if os.path.exists(fpath):
                    if app.test_win:
                        app.test_win.append_message_simple(f"❌ File '{fname}' already exists")
                    return

                try:
                    img.save(fpath)
                except Exception as e:
                    if app.test_win:
                        app.test_win.append_message_simple(f"❌ Failed to save image: {e}")
                    return

                if not os.path.exists(REFERENCE_FILE):
                    if app.test_win:
                        app.test_win.append_message_simple("⚠️ No reference image found")
                    return
                try:
                    ref_img_full = Image.open(REFERENCE_FILE).convert("RGB")
                except Exception as e:
                    if app.test_win:
                        app.test_win.append_message_simple(f"⚠️ Failed to open reference: {e}")
                    return

                try:
                    result_dict = compare_images_pipeline(img, ref_img_full, target_size=(300,300), do_align=True)
                except Exception as e:
                    if app.test_win:
                        app.test_win.append_message_simple(f"⚠️ Failed in pipeline: {e}")
                    return

                similarity = result_dict["combined_score"]

                min_val = app.sliders_win.min_var.get()
                max_val = app.sliders_win.max_var.get()

                passed = (min_val <= similarity <= max_val)
                result_text = "PASS" if passed else "FAIL"

                if app.test_win:
                    app.test_win.text_area.delete("1.0", tk.END)
                    app.test_win.text_area.config(bg="black")
                    if passed:
                        app.test_win.text_area.config(font=("Arial", 18), fg="green")
                        app.test_win.append_message_simple(f"PASS ✅ ({similarity:.2f}%)")
                    else:
                        app.test_win.text_area.config(font=("Arial", 17), fg="red")
                        app.test_win.append_message_simple(f"FAIL ❌ ({similarity:.2f}%)")

                if app.msg_viewer:
                    app.msg_viewer.append_message(
                        f"Result: {result_text}, Combined={similarity:.2f}%, "
                        f"SSIM={result_dict['ssim']:.2f}%, Diff={result_dict['diff_pct']:.2f}%, "
                        f"pHash={result_dict['phash_sim']:.2f}%, ORB={result_dict['orb_score']:.2f}%"
                    )

                try:
                    save_result_db(fname, similarity, result_text)
                except Exception:
                    pass

                def clear_result():
                    try:
                        if app.test_win:
                            app.test_win.text_area.delete("1.0", tk.END)
                    except Exception:
                        pass
                app.after(10000, clear_result)

        self.start_btn = DraggableButton(self.body, "Start Test", command=start_test_handler, init_pos=(pos_start[0], pos_start[1]), style=style)
        pos_exit = btn_positions.get("Exit", (160, 100))
        def exit_handler():
            self.master_app.on_close()
        self.exit_btn = DraggableButton(self.body, "Exit", command=exit_handler, init_pos=(pos_exit[0], pos_exit[1]), style=style)

    def _open_admin_panel(self):
        pw = simpledialog.askstring("Password", "Enter admin password:", show="*")
        if pw != self.settings.get("admin_password", DEFAULT_PASSWORD):
            messagebox.showerror("Error", "Wrong password!")
            return
        top = tk.Toplevel(self)
        top.title("Admin Panel")
        top.geometry("300x150")
        tk.Label(top, text="Change admin password").pack(pady=8)
        pw_var = tk.StringVar(value=self.settings.get("admin_password", DEFAULT_PASSWORD))
        entry = tk.Entry(top, textvariable=pw_var, show="*")
        entry.pack(pady=4)
        def save_pw():
            newpw = pw_var.get().strip()
            if not newpw:
                messagebox.showerror("Error", "Password cannot be empty")
                return
            self.settings["admin_password"] = newpw
            try:
                with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.settings, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            log_message("Admin password changed", self.master_app)
            messagebox.showinfo("Saved", "Password changed")
            top.destroy()
        tk.Button(top, text="Save", command=save_pw).pack(pady=6)

    def _start_move(self, ev):
        self._start_x = ev.x
        self._start_y = ev.y

    def _on_move(self, ev):
        dx = ev.x - self._start_x
        dy = ev.y - self._start_y
        info = self.place_info()
        self.place(x=int(float(info.get("x", 0)))+dx, y=int(float(info.get("y", 0)))+dy)

    def _start_resize(self, ev):
        info = self.place_info()
        self._start_w = int(info.get("width", 100))
        self._start_h = int(info.get("height", 100))
        self._start_x = ev.x_root
        self._start_y = ev.y_root

    def _on_resize(self, ev):
        dx = ev.x_root - self._start_x
        dy = ev.y_root - self._start_y
        new_w = max(150, self._start_w + dx)
        new_h = max(100, self._start_h + dy)
        info = self.place_info()
        self.place(x=int(float(info.get("x", 0))), y=int(float(info.get("y", 0))), width=new_w, height=new_h)

    def append_message(self, text):
        if self.text_area:
            self.text_area.insert(tk.END, text + "\\n")
            self.text_area.see(tk.END)

    def append_message_simple(self, text):
        if self.text_area:
            self.text_area.delete("1.0", tk.END)
            self.text_area.insert(tk.END, text)
            self.text_area.see(tk.END)

    def capture_frame(self):
        if self.video_cap is None:
            return None
        ret, frame = self.video_cap.read()
        if not ret:
            return None
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        return Image.fromarray(frame)

    def set_video_capture(self, cap):
        self.video_cap = cap

    def update_video_display(self, pil_img):
        if not self.video_label:
            return
        try:
            w = max(1, self.video_label.winfo_width())
            h = max(1, self.video_label.winfo_height())
            img = pil_img.resize((w, h))
            imgtk = ImageTk.PhotoImage(img)
            self.video_label.imgtk = imgtk
            self.video_label.config(image=imgtk)
        except Exception:
            pass

    def show_reference(self, pil_img):
        if not self.canvas:
            return
        # keep the original pil image for redraws during resize
        self._ref_img_pil = pil_img
        try:
            w = max(1, self.canvas.winfo_width())
            h = max(1, self.canvas.winfo_height())
            img = pil_img.resize((w, h))
            imgtk = ImageTk.PhotoImage(img)
            self.canvas.imgtk = imgtk
            self.canvas.delete("all")
            # place image at top-left
            self.canvas.create_image(0, 0, anchor="nw", image=imgtk)

            # اگر چندضلعی ذخیره شده بود دوباره بکش
            pts = self.settings.get("polygon_points", [])
            if pts and len(pts) > 1:
                for i in range(len(pts)):
                    x1, y1 = pts[i]
                    x2, y2 = pts[(i+1) % len(pts)]
                    self.canvas.create_line(x1, y1, x2, y2, fill="yellow", width=2)
                    self.canvas.create_oval(x1-3, y1-3, x1+3, y1+3, fill="red")

        except Exception:
            pass

    
    def start_polygon_mode(self):
        if not self.canvas:
            return
        self._polygon_points = []
        self._polygon_active = True
        # فعال کردن کلیک و دوبار کلیک
        self.canvas.bind("<Button-1>", self._on_polygon_click)
        self.canvas.bind("<Double-1>", self._finish_polygon)

    def _on_polygon_click(self, event):
        if not getattr(self, "_polygon_active", False):
            return
        x, y = event.x, event.y
        self._polygon_points.append((x, y))
        # رسم نقطه
        self.canvas.create_oval(x-3, y-3, x+3, y+3, fill="red")
        # رسم خط بین نقاط
        if len(self._polygon_points) > 1:
            self.canvas.create_line(self._polygon_points[-2], self._polygon_points[-1],
                                    fill="yellow", width=2)

    
    def _finish_polygon(self, event=None):
        if not getattr(self, "_polygon_active", False):
            return
        if len(self._polygon_points) > 2:
            # وصل کردن آخرین نقطه به اولین نقطه
            self.canvas.create_line(self._polygon_points[-1], self._polygon_points[0],
                                    fill="yellow", width=2)

            # ذخیره مختصات‌ها در تنظیمات
            self.settings["polygon_points"] = self._polygon_points
            try:
                with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.settings, f, ensure_ascii=False, indent=2)
            except Exception as e:
                log_message(f"Failed to save polygon points: {e}", self.master_app)

            # پیام موفقیت
            log_message("Polygon saved", self.master_app)
            if self.master_app.test_win:
                self.master_app.test_win.append_message_simple("Polygon ذخیره گردید ✅")
            # همچنین مختصات را در Message Viewer لاگ کن
            if self.master_app.msg_viewer:
                self.master_app.msg_viewer.append_message("Polygon points: " + str(self._polygon_points))

        self._polygon_active = False
        self.canvas.unbind("<Button-1>")
        self.canvas.unbind("<Double-1>")

        self.canvas.unbind("<Double-1>")

    def clear_polygon(self):
        if self.canvas:
            self.canvas.delete("all")
            # اگر تصویر reference وجود داشت دوباره بکشد
            if getattr(self, "_ref_img_pil", None):
                self.show_reference(self._ref_img_pil)

        # ریست کردن داده‌های چندضلعی
        self._polygon_points = []
        self._polygon_active = False
        self.settings["polygon_points"] = []   # <- پاک کردن داده‌های ذخیره‌شده

        # ذخیره در فایل تنظیمات
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(self.settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            log_message(f"Failed to reset polygon: {e}", self.master_app)

class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MDI Camera App v16")
        try:
            self.state("zoomed")
        except Exception:
            pass

        self.settings = {"windows": {}, "button_positions": {}, "admin_password": DEFAULT_PASSWORD, "sliders": {"min":0,"max":100}}
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                    self.settings = json.load(f)
            except Exception:
                pass

        self.client = tk.Frame(self, bg="#E5E7EB")
        self.client.pack(fill="both", expand=True)

        self.settings_win = MDIChild(self.client, "Settings", "settings", (10, 10, 400, 180), self.settings, with_buttons=True)
        self.sliders_win = MDIChild(self.client, "Sliders", "sliders", (420, 10, 260, 200), self.settings, with_sliders=True)
        self.live_win = MDIChild(self.client, "Live", "live", (300, 250, 640, 480), self.settings, with_video=True)
        self.ref_win = MDIChild(self.client, "Reference", "reference", (960, 250, 640, 480), self.settings, with_reference=True)

        # --- نمایش عکس مرجع و پلی‌گان ذخیره‌شده در شروع برنامه ---
        if os.path.exists(REFERENCE_FILE):
            try:
                img = Image.open(REFERENCE_FILE)
                self.ref_win.show_reference(img)
            except Exception as e:
                log_message(f"Failed to load reference on startup: {e}", self)
        self.test_win = MDIChild(self.client, "Test Result", "test_result", (10, 480, 300, 200), self.settings, with_test=True)
        self.msg_viewer = MDIChild(self.client, "Message Viewer", "msg", (660, 480, 640, 300), self.settings, with_text=True)

        # --- Lock/Unlock Buttons ---
        self.lock_state = False
        def toggle_lock():
            pw = simpledialog.askstring("Password", "Enter admin password:", show="*")
            if pw != self.settings.get("admin_password", DEFAULT_PASSWORD):
                messagebox.showerror("Error", "Wrong password!")
                return
            self.lock_state = not self.lock_state
            for child in (self.settings_win, self.sliders_win, self.live_win, self.ref_win, self.test_win):
                for w in child.body.winfo_children():
                    if isinstance(w, DraggableButton):
                        w.locked = self.lock_state
            if self.lock_state:
                log_message('🔒 Buttons locked', self)
            else:
                log_message('🔓 Buttons unlocked', self)

        lock_btn = tk.Button(self.client, text='Lock/Unlock Buttons', command=toggle_lock,
                             bg='purple', fg='white', font=('Arial', 10, 'bold'))
        lock_btn.place(x=20, y=750)

        # default device index (0=Webcam, 1=Capture Card)
        device_choice = self.settings.get("device", "Capture Card")
        self.device_index = 1 if device_choice == "Capture Card" else 0
        self.cap = None
        self._stop_event = threading.Event()
        self._start_capture()
        self.live_win.set_video_capture(self.cap)

        self.protocol("WM_DELETE_WINDOW", self.on_close)
        log_message("Application started", self)

    def _start_capture(self):
        try:
            # On Windows, use CAP_DSHOW to avoid long camera init delays
            self.cap = cv2.VideoCapture(self.device_index, cv2.CAP_DSHOW if os.name == "nt" else 0)
            if not self.cap.isOpened():
                log_message("Failed to open capture", self)
                self.cap = None
                return
        except Exception as e:
            log_message(f"Error opening capture: {e}", self)
            self.cap = None
            return
        t = threading.Thread(target=self._capture_loop, daemon=True)
        t.start()

    def _capture_loop(self):
        while not self._stop_event.is_set() and self.cap:
            ret, frame = self.cap.read()
            if not ret:
                continue
            try:
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                pil = Image.fromarray(frame_rgb)
                # schedule GUI update on main thread
                self.after(1, lambda p=pil: self._update_children_video(p))
            except Exception:
                pass
        if self.cap:
            try:
                self.cap.release()
            except Exception:
                pass

    def _update_children_video(self, pil_img):
        # update the live window only (other windows show captured reference)
        try:
            self.live_win.update_video_display(pil_img)
        except Exception:
            pass


    def change_device(self, choice):
        # Stop old capture
        self._stop_event.set()
        if self.cap:
            try:
                self.cap.release()
            except:
                pass
        self._stop_event = threading.Event()
        self.device_index = 1 if choice == "Capture Card" else 0
        self._start_capture()
        self.live_win.set_video_capture(self.cap)

    def on_close(self):
        self._stop_event.set()
        if self.cap:
            try:
                self.cap.release()
            except Exception:
                pass
        # save window geometry
        self.settings.setdefault("windows", {})
        for name, child in (("settings", self.settings_win),
                            ("sliders", self.sliders_win),
                            ("live", self.live_win),
                            ("reference", self.ref_win),
                            ("test_result", self.test_win),
                            ("msg", self.msg_viewer)):
            try:
                info = child.place_info()
                self.settings["windows"][name] = [int(float(info.get("x", 0))),
                                                  int(float(info.get("y", 0))),
                                                  int(float(info.get("width", 100))),
                                                  int(float(info.get("height", 100)))]
            except Exception:
                pass
        btn_pos = {}
        for btn in (getattr(self.settings_win, "ref_btn", None),
                    getattr(self.settings_win, "poly_btn", None),
                    getattr(self.settings_win, "reset_btn", None),
                    getattr(self.settings_win, "admin_btn", None),
                    getattr(self.sliders_win, "exit_btn", None),
                    getattr(self.settings_win, "device_label", None),
                    getattr(self.settings_win, "device_combo", None),
            getattr(self.sliders_win, "start_btn", None),
            getattr(self.sliders_win, "export_btn", None)):
            if btn is None:
                continue
            try:
                info = btn.place_info()
                # اگر text خالی بود، اسم اختصاصی بدهیم
                if btn["text"]:
                    key = btn["text"]
                else:
                    key = "Device Combo" if btn is getattr(self.settings_win, "device_combo", None) else "Device Label"
                btn_pos[key] = [int(float(info.get("x", 0))), int(float(info.get("y", 0)))]
            except Exception:
                pass
        self.settings["button_positions"] = btn_pos
        try:
            self.settings["sliders"] = {"min": int(self.sliders_win.min_var.get()), "max": int(self.sliders_win.max_var.get())}
        except Exception:
            pass
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(self.settings, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass


if __name__ == "__main__":
    app = MainApp()
    app.mainloop()
